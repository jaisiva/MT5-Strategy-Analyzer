"""
styles.py

Common Excel styles used throughout the MT5 Strategy Analyzer.

Provides reusable fonts, fills, borders, alignments and number formats.

Author:
    MT5 Strategy Analyzer

License:
    MIT
"""

from __future__ import annotations

from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)


class ExcelStyles:
    """
    Central repository of reusable Excel styles.
    """

    # -------------------------------------------------------------
    # Fonts
    # -------------------------------------------------------------

    TITLE_FONT = Font(
        name="Calibri",
        size=16,
        bold=True,
        color="1F1F1F",
    )

    HEADER_FONT = Font(
        name="Calibri",
        size=11,
        bold=True,
        color="FFFFFF",
    )

    NORMAL_FONT = Font(
        name="Calibri",
        size=11,
    )

    BOLD_FONT = Font(
        name="Calibri",
        size=11,
        bold=True,
    )

    # -------------------------------------------------------------
    # Fills
    # -------------------------------------------------------------

    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="4472C4",
    )

    SUCCESS_FILL = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    )

    LOSS_FILL = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    WARNING_FILL = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )

    INFO_FILL = PatternFill(
        fill_type="solid",
        fgColor="D9EAD3",
    )

    # -------------------------------------------------------------
    # Borders
    # -------------------------------------------------------------

    THIN = Side(
        border_style="thin",
        color="D9D9D9",
    )

    MEDIUM = Side(
        border_style="medium",
        color="808080",
    )

    THIN_BORDER = Border(
        left=THIN,
        right=THIN,
        top=THIN,
        bottom=THIN,
    )

    MEDIUM_BORDER = Border(
        left=MEDIUM,
        right=MEDIUM,
        top=MEDIUM,
        bottom=MEDIUM,
    )

    # -------------------------------------------------------------
    # Alignment
    # -------------------------------------------------------------

    CENTER = Alignment(
        horizontal="center",
        vertical="center",
    )

    LEFT = Alignment(
        horizontal="left",
        vertical="center",
    )

    RIGHT = Alignment(
        horizontal="right",
        vertical="center",
    )

    WRAP = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
    )

    # -------------------------------------------------------------
    # Number Formats
    # -------------------------------------------------------------

    CURRENCY = '$#,##0.00'

    PERCENT = '0.00%'

    DECIMAL = '0.00'

    INTEGER = '0'

    DATE = 'yyyy-mm-dd'

    DATETIME = 'yyyy-mm-dd hh:mm'

    TIME = 'hh:mm'

    # -------------------------------------------------------------
    # Helper Methods
    # -------------------------------------------------------------

    @classmethod
    def style_header(cls, cell) -> None:
        """
        Apply header style.
        """

        cell.font = cls.HEADER_FONT
        cell.fill = cls.HEADER_FILL
        cell.border = cls.THIN_BORDER
        cell.alignment = cls.CENTER

    @classmethod
    def style_title(cls, cell) -> None:
        """
        Apply worksheet title style.
        """

        cell.font = cls.TITLE_FONT
        cell.alignment = cls.LEFT

    @classmethod
    def style_currency(cls, cell) -> None:
        """
        Apply currency formatting.
        """

        cell.font = cls.NORMAL_FONT
        cell.border = cls.THIN_BORDER
        cell.alignment = cls.RIGHT
        cell.number_format = cls.CURRENCY

    @classmethod
    def style_percent(cls, cell) -> None:
        """
        Apply percentage formatting.
        """

        cell.font = cls.NORMAL_FONT
        cell.border = cls.THIN_BORDER
        cell.alignment = cls.RIGHT
        cell.number_format = cls.PERCENT

    @classmethod
    def style_decimal(cls, cell) -> None:
        """
        Apply decimal formatting.
        """

        cell.font = cls.NORMAL_FONT
        cell.border = cls.THIN_BORDER
        cell.alignment = cls.RIGHT
        cell.number_format = cls.DECIMAL

    @classmethod
    def style_integer(cls, cell) -> None:
        """
        Apply integer formatting.
        """

        cell.font = cls.NORMAL_FONT
        cell.border = cls.THIN_BORDER
        cell.alignment = cls.RIGHT
        cell.number_format = cls.INTEGER

    @classmethod
    def style_date(cls, cell) -> None:
        """
        Apply date formatting.
        """

        cell.font = cls.NORMAL_FONT
        cell.border = cls.THIN_BORDER
        cell.alignment = cls.CENTER
        cell.number_format = cls.DATE

    @classmethod
    def style_datetime(cls, cell) -> None:
        """
        Apply datetime formatting.
        """

        cell.font = cls.NORMAL_FONT
        cell.border = cls.THIN_BORDER
        cell.alignment = cls.CENTER
        cell.number_format = cls.DATETIME

    @classmethod
    def style_text(cls, cell) -> None:
        """
        Apply standard text formatting.
        """

        cell.font = cls.NORMAL_FONT
        cell.border = cls.THIN_BORDER
        cell.alignment = cls.LEFT
