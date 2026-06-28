"""
table_builder.py

Reusable Excel table writer.

Provides a consistent way to write formatted tables into worksheets.

Author:
    MT5 Strategy Analyzer

License:
    MIT
"""

from __future__ import annotations

from collections.abc import Mapping
from typing import Any

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from mt5_analyzer.reports.excel.styles import ExcelStyles


class TableBuilder:
    """
    Generic Excel table writer.
    """

    def __init__(
        self,
        sheet: Worksheet,
    ) -> None:

        self.sheet = sheet

    # --------------------------------------------------------------

    def write_table(
        self,
        *,
        title: str,
        data: list[Mapping[str, Any]],
        start_row: int = 1,
        start_col: int = 1,
        create_excel_table: bool = True,
        table_name: str | None = None,
    ) -> int:
        """
        Write a formatted table.

        Returns
        -------
        int
            Last row written.
        """

        if not data:

            return start_row

        # ----------------------------------------------------------
        # Title
        # ----------------------------------------------------------

        title_cell = self.sheet.cell(
            row=start_row,
            column=start_col,
        )

        title_cell.value = title

        ExcelStyles.style_title(
            title_cell
        )

        # ----------------------------------------------------------
        # Headers
        # ----------------------------------------------------------

        headers = list(
            data[0].keys()
        )

        header_row = start_row + 1

        for col, header in enumerate(
            headers,
            start=start_col,
        ):

            cell = self.sheet.cell(
                row=header_row,
                column=col,
            )

            cell.value = header

            ExcelStyles.style_header(
                cell
            )

        # ----------------------------------------------------------
        # Data
        # ----------------------------------------------------------

        current_row = header_row + 1

        for record in data:

            for col, header in enumerate(
                headers,
                start=start_col,
            ):

                value = record.get(
                    header
                )

                cell = self.sheet.cell(
                    row=current_row,
                    column=col,
                )

                cell.value = value

                self._style_cell(
                    cell,
                    value,
                )

            current_row += 1

        # ----------------------------------------------------------
        # Excel Table
        # ----------------------------------------------------------

        if create_excel_table:

            first = self.sheet.cell(
                row=header_row,
                column=start_col,
            ).coordinate

            last = self.sheet.cell(
                row=current_row - 1,
                column=start_col + len(headers) - 1,
            ).coordinate

            table = Table(

                displayName=table_name
                or f"Table_{header_row}",

                ref=f"{first}:{last}",

            )

            style = TableStyleInfo(

                name="TableStyleMedium2",

                showFirstColumn=False,

                showLastColumn=False,

                showRowStripes=True,

                showColumnStripes=False,

            )

            table.tableStyleInfo = style

            self.sheet.add_table(
                table
            )

        return current_row

    # --------------------------------------------------------------

    def _style_cell(
        self,
        cell,
        value,
    ) -> None:
        """
        Apply automatic formatting.
        """

        if value is None:

            ExcelStyles.style_text(
                cell
            )

            return

        if isinstance(
            value,
            bool,
        ):

            ExcelStyles.style_text(
                cell
            )

            return

        if isinstance(
            value,
            int,
        ):

            ExcelStyles.style_integer(
                cell
            )

            return

        if isinstance(
            value,
            float,
        ):

            ExcelStyles.style_decimal(
                cell
            )

            return

        try:

            import datetime

            if isinstance(
                value,
                datetime.datetime,
            ):

                ExcelStyles.style_datetime(
                    cell
                )

                return

            if isinstance(
                value,
                datetime.date,
            ):

                ExcelStyles.style_date(
                    cell
                )

                return

        except Exception:

            pass

        ExcelStyles.style_text(
            cell
        )
