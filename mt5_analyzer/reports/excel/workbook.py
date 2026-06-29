"""
workbook.py

Excel workbook helper for MT5 Strategy Analyzer.

Provides common workbook operations used by all report writers.

Responsibilities
----------------
- Create workbook
- Add/remove worksheets
- Get worksheet
- Activate worksheet
- Freeze panes
- Auto-fit columns
- Set worksheet zoom
- Save workbook

Author:
    MT5 Strategy Analyzer

License:
    MIT
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class ExcelWorkbook:
    """
    Wrapper around openpyxl Workbook.
    """

    def __init__(self) -> None:

        self.workbook = Workbook()

        # Remove default sheet.
        default = self.workbook.active
        self.workbook.remove(default)

    # -----------------------------------------------------------------

    def add_sheet(
        self,
        title: str,
    ) -> Worksheet:
        """
        Create a worksheet.
        """

        return self.workbook.create_sheet(
            title=title,
        )

    # -----------------------------------------------------------------

    def get_sheet(
        self,
        title: str,
    ) -> Worksheet:
        """
        Return worksheet by title.
        """

        if title not in self.workbook.sheetnames:

            raise KeyError(
                f"Worksheet '{title}' does not exist."
            )

        return self.workbook[title]

    # -----------------------------------------------------------------

    def activate_sheet(
        self,
        sheet: str | Worksheet,
    ) -> Worksheet:
        """
        Activate worksheet.
        """

        if isinstance(
            sheet,
            Worksheet,
        ):

            worksheet = sheet

        else:

            worksheet = self.get_sheet(
                sheet,
            )

        self.workbook.active = self.workbook.index(
            worksheet,
        )

        return worksheet

    # -----------------------------------------------------------------

    def remove_sheet(
        self,
        title: str,
    ) -> None:
        """
        Remove worksheet by title.
        """

        if title in self.workbook.sheetnames:

            self.workbook.remove(
                self.workbook[title]
            )

    # -----------------------------------------------------------------

    def freeze_header(
        self,
        sheet: Worksheet,
        row: int = 2,
    ) -> None:
        """
        Freeze top header rows.
        """

        sheet.freeze_panes = f"A{row}"

    # -----------------------------------------------------------------

    def auto_fit(
        self,
        sheet: Worksheet,
        min_width: int = 10,
        max_width: int = 50,
    ) -> None:
        """
        Auto-fit worksheet columns.
        """

        for column in sheet.columns:

            length = 0

            letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                try:

                    value = str(
                        cell.value
                    )

                except Exception:

                    value = ""

                length = max(
                    length,
                    len(value),
                )

            sheet.column_dimensions[
                letter
            ].width = max(

                min_width,

                min(
                    length + 2,
                    max_width,
                ),
            )

    # -----------------------------------------------------------------

    def set_zoom(
        self,
        sheet: Worksheet,
        zoom: int = 100,
    ) -> None:
        """
        Set worksheet zoom level.
        """

        sheet.sheet_view.zoomScale = zoom

    # -----------------------------------------------------------------

    def write_title(
        self,
        sheet: Worksheet,
        title: str,
        row: int = 1,
    ) -> None:
        """
        Write worksheet title.
        """

        cell = sheet.cell(
            row=row,
            column=1,
        )

        cell.value = title

        cell.font = Font(
            bold=True,
            size=14,
        )

    # -----------------------------------------------------------------

    def save(
        self,
        filename: str | Path,
    ) -> Path:
        """
        Save workbook.
        """

        path = Path(
            filename
        )

        path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        self.workbook.save(
            path
        )

        return path

    # -----------------------------------------------------------------

    @property
    def sheet_names(
        self,
    ) -> list[str]:
        """
        Workbook sheet names.
        """

        return self.workbook.sheetnames

    # -----------------------------------------------------------------

    def __len__(
        self,
    ) -> int:
        """
        Number of worksheets.
        """

        return len(
            self.workbook.sheetnames
        )
