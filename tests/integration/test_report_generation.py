"""
Integration tests for all report writers.
"""

from __future__ import annotations

from openpyxl import load_workbook

from mt5_analyzer.reports.excel.original_writer import (
    OriginalReportWriter,
)
from mt5_analyzer.reports.excel.lpattern_writer import (
    LPatternReportWriter,
)
from mt5_analyzer.reports.excel.drawdown_writer import (
    DrawdownReportWriter,
)
from mt5_analyzer.reports.excel.portfolio_writer import (
    PortfolioReportWriter,
)


class TestReportGeneration:

    def test_original_writer(
        self,
        output_dir,
        original_result,
    ):

        output = output_dir / "original.xlsx"

        OriginalReportWriter().write(
            original_result,
            output,
        )

        assert output.exists()

        wb = load_workbook(output)

        assert "Summary" in wb.sheetnames

        assert "Statistics" in wb.sheetnames

        assert "Trades" in wb.sheetnames

    # ------------------------------------------------------------

    def test_lpattern_writer(
        self,
        output_dir,
        lpattern_result,
    ):

        output = output_dir / "lpattern.xlsx"

        LPatternReportWriter().write(
            lpattern_result,
            output,
        )

        assert output.exists()

    # ------------------------------------------------------------

    def test_drawdown_writer(
        self,
        output_dir,
        drawdown_result,
    ):

        output = output_dir / "drawdown.xlsx"

        DrawdownReportWriter().write(
            drawdown_result,
            output,
        )

        assert output.exists()

    # ------------------------------------------------------------

    def test_portfolio_writer(
        self,
        output_dir,
        portfolio_result,
    ):

        output = output_dir / "portfolio.xlsx"

        PortfolioReportWriter().write(
            portfolio_result,
            output,
        )

        assert output.exists()
